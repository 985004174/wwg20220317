# /Users/kevin.wu/PycharmProjects python
# -*- coding: utf-8 -*-
# author： kevin.wu
# datetime： 2022/3/30 10:51 AM 
# ide： PyCharm
import openpyxl

from Api_keywords.api_key import ApiKey

# def read_excel():
file = openpyxl.load_workbook('../data/data1.xlsx')
# 创建workbook对象，列表对象

# 定位sheet页
sheet = file['Sheet1']

# 写入单元格数据
sheet.cell(row=1,column=1).value='测试写入数据'

file.save('../data/data1.xlsx')

# 读取单元格数据
res=sheet.cell(1,2).value

res2 = sheet.cell(2,3).value


# print(sheet.max_row,sheet.min_row,sheet.max_column,sheet.min_column)


all_content = []
for i in range(1, sheet.max_row+1):
    row_content = []
    for j in range(1, sheet.max_column+1):
        value = sheet.cell(row=i, column=j).value
        row_content.append(value)
    all_content.append(row_content)
    print(row_content)
    print('*******************************************************')
# print(all_content)





#
# all_row = []
# for row in sheet.rows:       # 按行获取单元格(Cell对象) - 生成器
#     ev_row = []
#     for cell in row:
#         value = cell.value
#         ev_row.append(value)
#     all_row.append(ev_row)
#     print(ev_row)
#     print('*******************************************************')
#     print(all_row)



# 循环读取sheet页的值






























#     ak=ApiKey()
#     title_list=[]
#     r_method_list=[]
#
#     for value in sheet.values:
#         # print(value)
#
#         if value[0] is int:
#             print('====用例开始执行====')
#
#             dict={
#                 'url': value[1] + value[2],
#                 'hd': eval(value[4]),
#                 'data':eval(value[5])
#             }
#             print(dict)
#         else:
#             print('====用例暂未执行=====')
#
# read_excel()

#
#
#

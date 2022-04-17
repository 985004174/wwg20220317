# /Users/kevin.wu/PycharmProjects python
# -*- coding: utf-8 -*-
# author： kevin.wu
# datetime： 2022/4/11 5:26 PM 
# ide： PyCharm
import os
from datetime import datetime
import eth_account
import openpyxl
from log.logger_config import test_log

log = test_log()


def write(x):
    # 创建时间戳

    times = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    workbook = openpyxl.load_workbook('../data/eth_account/eth_address.xlsx')
    sheet = workbook['eth']
    for i in range(2, int(x)):
        account = eth_account.Account.create()

        # 输出地址
        address = account.address

        # 输出私钥
        private_key = account.key.hex()


        sheet.cell(i, 1).value = i - 1
        sheet.cell(i, 2).value = address
        sheet.cell(i, 3).value = private_key
        sheet.cell(i, 4).value = None

    workbook.save('../data/eth_account/eth_address-{}.xlsx'.format(times))

    if sheet.cell(2,2) is not None:
        print('当前文件已经生成')
    else:
        e = '985004174@qq.com'
        print('文件生成失败，请联系管理员:{}'.format(e))
    workbook.close()

if __name__ == '__main__':

    try:
        x=input('请输入要生成的数量：')
        write(x)
    except Exception as e:
        print('输入错误:{},请输入数字'.format(e))
        x = input('请输入要生成的数量：')
        write(x)





